import pandas as pd
import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

class PerformanceData:
    header = ["Number of Threads", "Array Length", "Time", "Serial Time", "Speedup", "Efficiency", "karpFlatt"]
    def __init__(self, numThreads, arrayLength, time, serialTime):
        
        self.numThreads = numThreads
        self.arrayLength = arrayLength
        self.time = time
        self.serialTime = serialTime
    
        self.speedup = self.serialTime / self.time
        self.efficiency = self.time / self.numThreads

        if numThreads == 1:
            self.karpFlatt = 1
        else:
            self.karpFlatt = ((1/self.speedup) - (1/numThreads)) / (1-(1/numThreads))
    
    
    def __str__(self):
        return (f'numThreads={self.numThreads}, arrayLength={self.arrayLength}, '
                f'time={self.time}, serialTime={self.serialTime}, speedup={self.speedup:.2f}, '
                f'efficiency={self.efficiency:.2f}, KarpFlatt={self.karpFlatt:.2f}')
            
def exportToExcel(objects):
    data = {
        'numThreads': [],
        'arrayLength': [],
        'time': [],
        'serialTime': [],
        'speedup': [],
        'efficiency': [],
        'KarpFlatt': []
    }   

    for obj in objects:
        data['numThreads'].append(obj.numThreads)
        data['arrayLength'].append(obj.arrayLength)
        data['time'].append(obj.time)
        data['serialTime'].append(obj.serialTime)
        data['speedup'].append(obj.speedup)
        data['efficiency'].append(obj.efficiency)
        data['KarpFlatt'].append(obj.karpFlatt)

    df = pd.DataFrame(data)

    # Export the DataFrame to an Excel file
    file_name = 'performance_data.xlsx'
    sheetName = 'sheet 1'
    

    wb = Workbook()
    ws = wb.active
    cs = wb.create_chartsheet()

    ws.append(PerformanceData.header)
    
    for obj in objects:
        ws.append([obj.numThreads, obj.arrayLength, obj.time, obj.serialTime, obj.speedup, obj.efficiency, obj.karpFlatt])
        
    lastRow = ws.max_row
    performanceTable = Table(displayName="PerformanceData", ref=f"A1:G{lastRow}")

    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    performanceTable.tableStyleInfo = style
    
    ws.add_table(performanceTable)

    
    addAvgTable(ws, df, 'speedup', 9, "AverageSpeedup")
    addAvgTable(ws, df, 'efficiency', 12, "AverageEfficiency")
    addAvgTable(ws, df, 'KarpFlatt', 15, "AverageKarpFlatt")

    create_line_chart(ws, "Average Speedup", 9, 1, 9, 'AverageSpeedup', "I15")
    create_line_chart(ws, "Average Efficiency", 12, 1, 9, 'AverageEfficiency', "I30")
    create_line_chart(ws, "Average Karp-Flatt", 15, 1, 9, 'AverageKarpFlatt', "I45")
    
    
    wb.save(file_name)

def addAvgTable(ws, df, avgColumn, startingCol, title):
    avg_per_thread = df.groupby('numThreads')[avgColumn].mean().reset_index()
    avg_per_thread.columns = ['Number of Threads', title]

    avg_start_row = 1
    avg_start_col = startingCol
    ws.cell(row=avg_start_row, column=avg_start_col, value="Number of Threads")
    ws.cell(row=avg_start_row, column=avg_start_col+1, value=title)
    

    for index, row in avg_per_thread.iterrows():
        ws.cell(row=avg_start_row + index + 1, column=avg_start_col, value=row['Number of Threads'])
        ws.cell(row=avg_start_row + index + 1, column=avg_start_col+1, value=row[title])
    
    avgTable_ref = f"{chr(64 + avg_start_col)}{avg_start_row}:{chr(64 + avg_start_col + 1)}{12 + avg_start_row}"
    
    avgTable = Table(displayName=title, ref=avgTable_ref)

    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    avgTable.tableStyleInfo = style
    ws.add_table(avgTable)

def create_line_chart(ws, title, data_start_col, data_start_row, num_threads_col, avg_col, chart_position):
    # Create a LineChart
    line_chart = LineChart()
    line_chart.title = f"{title} Chart"
    line_chart.style = 13
    line_chart.x_axis.title = 'Number of Threads'
    line_chart.y_axis.title = title  # Set Y-axis title to the chart title for clarity
    
    # Reference the data for the average column
    data = Reference(ws,
                     min_col=data_start_col + 1,  # Reference the average data column
                     min_row=data_start_row + 1,  # Start from the row after the title
                     max_col=data_start_col + 1,
                     max_row=ws.max_row)

    # Reference the x-axis (Number of Threads)
    categories = Reference(ws,
                           min_col=num_threads_col,
                           min_row=data_start_row + 1,
                           max_row=ws.max_row)

    # Add the data and categories to the chart
    line_chart.add_data(data, titles_from_data=False)  # Set titles_from_data to False for single line
    line_chart.set_categories(categories)
    line_chart.legend = None
    

    # Position the chart
    ws.add_chart(line_chart, chart_position)

def csv_to_objects(file_path):
    df = pd.read_csv(file_path)
    
    
    objects = []

    for index, row in df.iterrows():
        obj = PerformanceData(
            numThreads=row['numThreads'],
            arrayLength=row[' arrayLength'],
            time=row[' time'],
            serialTime=row[' serialTime']
        )
        objects.append(obj)
    return objects


def printObjects(objects):
    for obj in objects:
        print(obj)
        


def main():
    parser = argparse.ArgumentParser(description="Load CSV into a list of objects")
    parser.add_argument('file', help="Path to the CSV file")

    args = parser.parse_args()

    objects = csv_to_objects(args.file)

    # printObjects(objects)
    exportToExcel(objects)


if __name__ == '__main__':
    main()

